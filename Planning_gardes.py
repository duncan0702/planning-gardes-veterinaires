from ortools.sat.python import cp_model
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os

class VetSchedulerHistory:
    """Gestion de l'historique des gardes"""
    
    def __init__(self, filepath='historique_gardes.json'):
        self.filepath = filepath
        self.history = self.load()
    
    def load(self):
        """Charge l'historique depuis le fichier JSON"""
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠️ Erreur lors du chargement de l'historique: {e}")
                return {}
        return {}
    
    def save(self):
        """Sauvegarde l'historique dans le fichier JSON"""
        try:
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
            print(f"✅ Historique sauvegardé dans {self.filepath}")
        except Exception as e:
            print(f"❌ Erreur lors de la sauvegarde de l'historique: {e}")
    
    def add_schedule(self, schedule, period_name=None):
        """Ajoute un planning à l'historique"""
        if not schedule:
            return
        
        if period_name is None:
            # Générer un nom basé sur les dates
            first_date = schedule[0]['date']
            last_date = schedule[-1]['date']
            period_name = f"{first_date}_to_{last_date}"
        
        stats = defaultdict(lambda: {
            'premier_semaine': 0, 
            'premier_weekend': 0, 
            'rappelable_semaine': 0,
            'deuxieme_weekend': 0
        })
        
        for entry in schedule:
            date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
            is_weekend = date_obj.weekday() >= 5
            
            if entry['premier']:
                if is_weekend:
                    stats[entry['premier']]['premier_weekend'] += 1
                else:
                    stats[entry['premier']]['premier_semaine'] += 1
            
            if entry['rappelable']:
                stats[entry['rappelable']]['rappelable_semaine'] += 1
            
            if entry['deuxieme']:
                stats[entry['deuxieme']]['deuxieme_weekend'] += 1
        
        # Convertir en format sérialisable
        self.history[period_name] = {
            'date_debut': schedule[0]['date'],
            'date_fin': schedule[-1]['date'],
            'stats': {vet: dict(s) for vet, s in stats.items()}
        }
        
        self.save()
        print(f"✅ Planning ajouté à l'historique: {period_name}")
    
    def get_cumulative_stats(self, vet_names):
        """Calcule les statistiques cumulées pour tous les vétérinaires"""
        cumul = {vet: {
            'premier_semaine': 0,
            'premier_weekend': 0, 
            'rappelable_semaine': 0,
            'deuxieme_weekend': 0
        } for vet in vet_names}
        
        for period_name, period_data in self.history.items():
            for vet, stats in period_data['stats'].items():
                if vet in cumul:
                    for key, value in stats.items():
                        cumul[vet][key] += value
        
        return cumul
    
    def print_history(self):
        """Affiche l'historique"""
        if not self.history:
            print("📋 Aucun historique disponible")
            return
        
        print("\n" + "="*80)
        print("HISTORIQUE DES PLANNINGS")
        print("="*80)
        
        for period_name, period_data in self.history.items():
            print(f"\n📅 Période: {period_data['date_debut']} → {period_data['date_fin']}")
            print(f"{'Vétérinaire':<20} {'1er sem':<10} {'1er WE':<10} {'Rappel':<10} {'2ème WE':<10}")
            print("-" * 70)
            
            for vet, stats in sorted(period_data['stats'].items()):
                print(f"{vet:<20} {stats['premier_semaine']:<10} "
                      f"{stats['premier_weekend']//2:<10} "
                      f"{stats['rappelable_semaine']:<10} "
                      f"{stats['deuxieme_weekend']//2:<10}")
        
        print("="*80)
    
    def clear(self):
        """Efface tout l'historique"""
        self.history = {}
        self.save()
        print("✅ Historique effacé")


class VetSchedulerConfig:
    """Configuration externalisée pour le planificateur"""
    
    def __init__(self):
        self.groupe_A = ['Dr. Timoty', 'Dr. Laura', 'Dr. Lauranne', 'Dr. Sarah', 'Dr. Malaurie']
        self.groupe_B = ['Dr. Julien', 'Dr. Isaure', 'Dr. Maxime', 'Dr. Nicolas', 'Dr. Mélanie', 'Dr. Olivier', 'Dr. Dorra']
        
        self.vets_speciaux = {
            'Dr. Olivier': {
                'jamais_premier': True,
                'jamais_deuxieme': True,
                'jamais_weekend': True,
                'max_rappelable_2_semaines': 1
            },
            'Dr. Julien': {
                'peut_garde_veille_repos': True
            },
            'Dr. Laura': {
                'peut_garde_weekend_malgre_repos_lundi': True
            },
            'Dr. Dorra': {
                'jamais_lundi': True,
                'max_rappelable_par_semaine': 1,
                'exclus_compatibilite_binome': True  # Option 4
            }
        }
        
        self.contraintes = {
            'max_premier_semaine': 1,
            'max_rappelable_semaine': 2,
            'ecart_equilibrage_premier': 2,
            'ecart_equilibrage_rappelable': 2,
            'ecart_equilibrage_deuxieme': 2,
            'espacement_weekends_jours': 14,
            'max_sequences_consecutives_rappelable': 1
        }
    
    @classmethod
    def from_json(cls, filepath):
        """Charge la configuration depuis un fichier JSON"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        config = cls()
        if 'groupe_A' in data:
            config.groupe_A = data['groupe_A']
        if 'groupe_B' in data:
            config.groupe_B = data['groupe_B']
        if 'vets_speciaux' in data:
            config.vets_speciaux = data['vets_speciaux']
        if 'contraintes' in data:
            config.contraintes.update(data['contraintes'])
        
        return config
    
    def to_json(self, filepath):
        """Sauvegarde la configuration vers un fichier JSON"""
        data = {
            'groupe_A': self.groupe_A,
            'groupe_B': self.groupe_B,
            'vets_speciaux': self.vets_speciaux,
            'contraintes': self.contraintes
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)


class VetScheduler:
    def __init__(self, start_date, end_date, veterinaires, config=None, history=None):
        """
        Planificateur de gardes vétérinaires
        
        Args:
            start_date: Date de début (format 'YYYY-MM-DD')
            end_date: Date de fin (format 'YYYY-MM-DD')
            veterinaires: dict avec structure vétérinaires
            config: VetSchedulerConfig (optionnel)
            history: VetSchedulerHistory (optionnel) - pour équilibrage sur plusieurs mois
        """
        # Validation des données d'entrée
        self.validate_inputs(start_date, end_date, veterinaires)
        
        self.start_date = datetime.strptime(start_date, '%Y-%m-%d')
        self.end_date = datetime.strptime(end_date, '%Y-%m-%d')
        self.veterinaires = veterinaires
        self.vet_names = list(veterinaires.keys())
        self.n_vets = len(self.vet_names)
        
        # Configuration
        self.config = config if config else VetSchedulerConfig()
        
        # Historique
        self.history = history if history else VetSchedulerHistory()
        self.historical_stats = self.history.get_cumulative_stats(self.vet_names)
        
        # Générer la liste des jours
        self.dates = []
        current = self.start_date
        while current <= self.end_date:
            self.dates.append(current)
            current += timedelta(days=1)
        
        self.n_days = len(self.dates)
        
        # Identifier les jours de semaine, semaines civiles et week-ends
        self.weekdays = []  # Tous les jours lundi-vendredi
        self.weeks = []  # Semaines civiles (lundi-vendredi groupés)
        self.weekends = []  # Week-ends (samedi-dimanche groupés)
        
        i = 0
        current_week = []
        while i < self.n_days:
            date = self.dates[i]
            if date.weekday() < 5:  # Lundi à vendredi
                self.weekdays.append(i)
                current_week.append(i)
                
                # Si vendredi ou dernier jour, terminer la semaine
                if date.weekday() == 4 or i == self.n_days - 1:
                    if current_week:
                        self.weeks.append(current_week)
                        current_week = []
                i += 1
            else:  # Week-end
                weekend_days = []
                if i < self.n_days and self.dates[i].weekday() == 5:
                    weekend_days.append(i)
                    i += 1
                if i < self.n_days and self.dates[i].weekday() == 6:
                    weekend_days.append(i)
                    i += 1
                if weekend_days:
                    self.weekends.append(weekend_days)
        
        print(f"\n📊 Période: {self.start_date.strftime('%Y-%m-%d')} -> {self.end_date.strftime('%Y-%m-%d')}")
        print(f"   Total jours: {self.n_days}")
        print(f"   Jours de semaine: {len(self.weekdays)}")
        print(f"   Week-ends: {len(self.weekends)}")
        print(f"   Vétérinaires: {self.n_vets}")
        
        self.model = cp_model.CpModel()
        self.create_variables()
    
    def validate_inputs(self, start_date, end_date, veterinaires):
        """Valide les données d'entrée"""
        # Validation des dates
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError as e:
            raise ValueError(f"Format de date invalide (doit être YYYY-MM-DD): {e}")
        
        if start > end:
            raise ValueError(f"La date de début ({start_date}) doit être avant la date de fin ({end_date})")
        
        if (end - start).days > 365:
            raise ValueError("La période ne peut pas dépasser 1 an")
        
        # Validation des vétérinaires
        if not veterinaires:
            raise ValueError("La liste des vétérinaires ne peut pas être vide")
        
        if len(veterinaires) < 3:
            raise ValueError("Il faut au moins 3 vétérinaires pour générer un planning")
        
        for vet_name, info in veterinaires.items():
            # Vérifier la structure
            if not isinstance(info, dict):
                raise ValueError(f"{vet_name}: Les informations doivent être un dictionnaire")
            
            if 'jour_repos' not in info:
                raise ValueError(f"{vet_name}: Le champ 'jour_repos' est obligatoire")
            
            # Valider jour_repos
            jours_repos = info['jour_repos']
            if isinstance(jours_repos, int):
                if not 0 <= jours_repos <= 6:
                    raise ValueError(f"{vet_name}: jour_repos doit être entre 0 (lundi) et 6 (dimanche), reçu {jours_repos}")
            elif isinstance(jours_repos, list):
                if not all(isinstance(j, int) and 0 <= j <= 6 for j in jours_repos):
                    raise ValueError(f"{vet_name}: Tous les jours de repos doivent être entre 0 et 6")
                if len(jours_repos) > 5:
                    raise ValueError(f"{vet_name}: Trop de jours de repos ({len(jours_repos)}), maximum 5")
            else:
                raise ValueError(f"{vet_name}: jour_repos doit être un entier ou une liste d'entiers")
            
            # Valider congés
            if 'conges' in info:
                if not isinstance(info['conges'], list):
                    raise ValueError(f"{vet_name}: Le champ 'conges' doit être une liste")
                
                for conge in info['conges']:
                    try:
                        datetime.strptime(conge, '%Y-%m-%d')
                    except ValueError:
                        raise ValueError(f"{vet_name}: Date de congé invalide '{conge}' (format requis: YYYY-MM-DD)")
        
        print("✅ Validation des données réussie")
    
    def diagnose_schedule(self, schedule):
        """Diagnostic automatique du planning généré"""
        if not schedule:
            return {"status": "error", "message": "Aucun planning à diagnostiquer"}
        
        violations = []
        warnings = []
        stats = defaultdict(lambda: {'premier_semaine': 0, 'premier_weekend': 0, 
                                     'rappelable_semaine': 0, 'deuxieme_weekend': 0})
        
        print("\n" + "="*80)
        print("DIAGNOSTIC AUTOMATIQUE DU PLANNING")
        print("="*80)
        
        # Vérifier chaque jour
        for idx, entry in enumerate(schedule):
            date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
            is_weekend = date_obj.weekday() >= 5
            
            # Vérifier qu'il y a bien un premier et un second
            if not entry['premier']:
                violations.append(f"❌ {entry['date']}: Aucun premier de garde")
            
            if is_weekend:
                if not entry['deuxieme']:
                    violations.append(f"❌ {entry['date']}: Aucun 2ème de garde (week-end)")
                if entry['rappelable']:
                    violations.append(f"❌ {entry['date']}: Rappelable présent le week-end (devrait être 2ème)")
            else:
                if not entry['rappelable']:
                    violations.append(f"❌ {entry['date']}: Aucun rappelable (semaine)")
                if entry['deuxieme']:
                    violations.append(f"❌ {entry['date']}: 2ème de garde en semaine (devrait être rappelable)")
            
            # Vérifier que premier et second sont différents
            if entry['premier'] and entry['rappelable'] and entry['premier'] == entry['rappelable']:
                violations.append(f"❌ {entry['date']}: Même personne en premier et rappelable")
            if entry['premier'] and entry['deuxieme'] and entry['premier'] == entry['deuxieme']:
                violations.append(f"❌ {entry['date']}: Même personne en premier et 2ème de garde")
            
            # Accumuler les statistiques
            if entry['premier']:
                if is_weekend:
                    stats[entry['premier']]['premier_weekend'] += 1
                else:
                    stats[entry['premier']]['premier_semaine'] += 1
            
            if entry['rappelable']:
                stats[entry['rappelable']]['rappelable_semaine'] += 1
            
            if entry['deuxieme']:
                stats[entry['deuxieme']]['deuxieme_weekend'] += 1
        
        # Vérifier l'équilibrage
        premiers_semaine = [s['premier_semaine'] for s in stats.values()]
        premiers_weekend = [s['premier_weekend'] // 2 for s in stats.values()]  # Diviser par 2 (sam+dim)
        rappelables = [s['rappelable_semaine'] for s in stats.values()]
        deuxiemes = [s['deuxieme_weekend'] // 2 for s in stats.values()]
        
        if premiers_semaine:
            ecart_premier_semaine = max(premiers_semaine) - min(premiers_semaine)
            if ecart_premier_semaine > self.config.contraintes['ecart_equilibrage_premier']:
                warnings.append(f"⚠️ Équilibrage premier semaine: écart de {ecart_premier_semaine} (max recommandé: {self.config.contraintes['ecart_equilibrage_premier']})")
        
        if premiers_weekend and max(premiers_weekend) > 0:
            ecart_premier_weekend = max(premiers_weekend) - min(premiers_weekend)
            if ecart_premier_weekend > self.config.contraintes['ecart_equilibrage_premier']:
                warnings.append(f"⚠️ Équilibrage premier week-end: écart de {ecart_premier_weekend}")
        
        if rappelables:
            ecart_rappelable = max(rappelables) - min(rappelables)
            if ecart_rappelable > self.config.contraintes['ecart_equilibrage_rappelable']:
                warnings.append(f"⚠️ Équilibrage rappelable: écart de {ecart_rappelable} (max recommandé: {self.config.contraintes['ecart_equilibrage_rappelable']})")
        
        # Afficher les résultats
        if violations:
            print(f"\n🔴 VIOLATIONS DÉTECTÉES ({len(violations)}):")
            for v in violations[:10]:  # Limiter à 10 pour lisibilité
                print(f"  {v}")
            if len(violations) > 10:
                print(f"  ... et {len(violations) - 10} autres violations")
        else:
            print("\n✅ Aucune violation détectée")
        
        if warnings:
            print(f"\n⚠️ AVERTISSEMENTS ({len(warnings)}):")
            for w in warnings:
                print(f"  {w}")
        else:
            print("\n✅ Aucun avertissement")
        
        # Statistiques de charge
        print("\n📊 STATISTIQUES DE CHARGE (période actuelle):")
        print(f"{'Vétérinaire':<20} {'1er sem':<10} {'1er WE':<10} {'Rappel':<10} {'2ème WE':<10} {'Total':<10}")
        print("-" * 80)
        
        for vet_name in sorted(stats.keys()):
            s = stats[vet_name]
            total = s['premier_semaine'] + s['premier_weekend'] // 2 + s['rappelable_semaine'] + s['deuxieme_weekend'] // 2
            print(f"{vet_name:<20} {s['premier_semaine']:<10} {s['premier_weekend']//2:<10} "
                  f"{s['rappelable_semaine']:<10} {s['deuxieme_weekend']//2:<10} {total:<10}")
        
        # Afficher les statistiques cumulées avec l'historique
        if any(sum(h.values()) > 0 for h in self.historical_stats.values()):
            print("\n📚 STATISTIQUES CUMULÉES (historique + période actuelle):")
            print(f"{'Vétérinaire':<20} {'1er sem':<10} {'1er WE':<10} {'Rappel':<10} {'2ème WE':<10} {'Total':<10}")
            print("-" * 80)
            
            for vet_name in sorted(stats.keys()):
                s = stats[vet_name]
                h = self.historical_stats.get(vet_name, {})
                
                total_1er_sem = s['premier_semaine'] + h.get('premier_semaine', 0)
                total_1er_we = (s['premier_weekend'] + h.get('premier_weekend', 0)) // 2
                total_rappel = s['rappelable_semaine'] + h.get('rappelable_semaine', 0)
                total_2eme_we = (s['deuxieme_weekend'] + h.get('deuxieme_weekend', 0)) // 2
                total_global = total_1er_sem + total_1er_we + total_rappel + total_2eme_we
                
                print(f"{vet_name:<20} {total_1er_sem:<10} {total_1er_we:<10} "
                      f"{total_rappel:<10} {total_2eme_we:<10} {total_global:<10}")
        
        print("="*80)
        
        return {
            "status": "success" if not violations else "error",
            "violations": violations,
            "warnings": warnings,
            "stats": dict(stats)
        }
        
    def create_variables(self):
        """Crée les variables de décision"""
        self.premier = {}
        for v in range(self.n_vets):
            for d in range(self.n_days):
                self.premier[(v, d)] = self.model.NewBoolVar(f'premier_v{v}_d{d}')
        
        self.rappelable = {}
        for v in range(self.n_vets):
            for d in range(self.n_days):
                self.rappelable[(v, d)] = self.model.NewBoolVar(f'rappelable_v{v}_d{d}')
        
        # NOUVEAU: 2ème de garde pour les week-ends uniquement
        self.deuxieme = {}
        for v in range(self.n_vets):
            for d in range(self.n_days):
                self.deuxieme[(v, d)] = self.model.NewBoolVar(f'deuxieme_v{v}_d{d}')

    def add_constraints(self):
        """Ajoute toutes les contraintes"""
        
        print("\n--- AJOUT DES CONTRAINTES ---")
        
        # 1. Chaque jour de semaine: exactement 1 premier et 1 rappelable différent
        print("✓ Contrainte 1: 1 premier + 1 rappelable par jour de semaine")
        for d in self.weekdays:
            self.model.Add(sum(self.premier[(v, d)] for v in range(self.n_vets)) == 1)
            self.model.Add(sum(self.rappelable[(v, d)] for v in range(self.n_vets)) == 1)
            for v in range(self.n_vets):
                self.model.Add(self.premier[(v, d)] + self.rappelable[(v, d)] <= 1)
                # Pas de 2ème de garde en semaine
                self.model.Add(self.deuxieme[(v, d)] == 0)
        
        # 2. MODIFIÉ: Chaque week-end: mêmes vétérinaires samedi et dimanche (1er + 2ème de garde)
        print("✓ Contrainte 2: Mêmes personnes samedi-dimanche (1er + 2ème de garde)")
        for weekend_days in self.weekends:
            if len(weekend_days) == 2:
                sam, dim = weekend_days
                # 1 premier et 1 deuxième de garde (pas rappelable)
                self.model.Add(sum(self.premier[(v, sam)] for v in range(self.n_vets)) == 1)
                self.model.Add(sum(self.deuxieme[(v, sam)] for v in range(self.n_vets)) == 1)
                
                # Pas de rappelable le week-end
                for v in range(self.n_vets):
                    self.model.Add(self.rappelable[(v, sam)] == 0)
                    self.model.Add(self.rappelable[(v, dim)] == 0)
                
                for v in range(self.n_vets):
                    # Mêmes vétérinaires samedi et dimanche
                    self.model.Add(self.premier[(v, sam)] == self.premier[(v, dim)])
                    self.model.Add(self.deuxieme[(v, sam)] == self.deuxieme[(v, dim)])
                    # Premier et 2ème différents
                    self.model.Add(self.premier[(v, sam)] + self.deuxieme[(v, sam)] <= 1)
        
        # 3. Maximum 1 garde premier par semaine civile
        print(f"✓ Contrainte 3: Max {self.config.contraintes['max_premier_semaine']} garde premier par semaine civile")
        for v in range(self.n_vets):
            for week_days in self.weeks:
                self.model.Add(sum(self.premier[(v, d)] for d in week_days) <= self.config.contraintes['max_premier_semaine'])
        
        # 4. Maximum 2 rappelable par semaine civile (sauf Dr. Olivier: max 1 toutes les 2 semaines)
        print("✓ Contrainte 4: Max 2 rappelable par semaine civile")
        for v in range(self.n_vets):
            vet_name = self.vet_names[v]
            
            if vet_name == 'Dr. Olivier':
                # Dr. Olivier: max 1 rappelable sur 2 semaines consécutives
                for i in range(len(self.weeks) - 1):
                    week1 = self.weeks[i]
                    week2 = self.weeks[i + 1]
                    combined_weeks = week1 + week2
                    self.model.Add(sum(self.rappelable[(v, d)] for d in combined_weeks) <= 1)
            elif vet_name == 'Dr. Dorra':
                # Dr. Dorra: max 1 rappelable par semaine civile
                for week_days in self.weeks:
                    self.model.Add(sum(self.rappelable[(v, d)] for d in week_days) <= 1)
            else:
                # Autres vétérinaires: max 2 par semaine civile
                for week_days in self.weeks:
                    self.model.Add(sum(self.rappelable[(v, d)] for d in week_days) <= 2)
        
        # 5. Repos obligatoire le lendemain d'une garde premier EN SEMAINE
        print("✓ Contrainte 5: Repos obligatoire après garde premier en semaine")
        for v in range(self.n_vets):
            for d in self.weekdays:
                if d + 1 < self.n_days:
                    self.model.AddImplication(self.premier[(v, d)], self.premier[(v, d + 1)].Not())
                    self.model.AddImplication(self.premier[(v, d)], self.rappelable[(v, d + 1)].Not())
        
        # 5bis. Repos le lundi après un week-end en premier (SAUF Dr. Laura si lundi = jour de repos)
        print("✓ Contrainte 5bis: Repos le lundi après week-end en premier (exception Dr. Laura)")
        for v in range(self.n_vets):
            vet_name = self.vet_names[v]
            vet_info = self.veterinaires[vet_name]
            
            # Récupérer les jours de repos
            jours_repos = vet_info['jour_repos']
            if isinstance(jours_repos, int):
                jours_repos = [jours_repos]
            
            # Récupérer les congés
            conges = [datetime.strptime(c, '%Y-%m-%d') for c in vet_info.get('conges', [])]
            
            for weekend_days in self.weekends:
                if len(weekend_days) == 2:
                    dim = weekend_days[1]
                    if dim + 1 < self.n_days and self.dates[dim + 1].weekday() == 0:
                        lundi = dim + 1
                        lundi_date = self.dates[lundi]
                        
                        # EXCEPTION pour Dr. Laura: si le lundi est son jour de repos ET qu'elle n'est pas en congé
                        if vet_name == 'Dr. Laura' and 0 in jours_repos and lundi_date not in conges:
                            # Dr. Laura peut faire garde le week-end même si elle a repos le lundi
                            # On ne bloque PAS la garde week-end
                            pass
                        else:
                            # Pour tous les autres (ou Dr. Laura en congé le lundi)
                            self.model.AddImplication(
                                self.premier[(v, weekend_days[0])],
                                self.premier[(v, lundi)].Not()
                            )
                            self.model.AddImplication(
                                self.premier[(v, weekend_days[0])],
                                self.rappelable[(v, lundi)].Not()
                            )
        
        # 6. Max 1 séquence de 2 jours consécutifs en rappelable par semaine
        print("✓ Contrainte 6: Max 1 séquence de 2 jours consécutifs rappelable par semaine")
        for v in range(self.n_vets):
            for week_days in self.weeks:
                consecutive_pairs = []
                for i in range(len(week_days) - 1):
                    d1, d2 = week_days[i], week_days[i + 1]
                    if d2 == d1 + 1:
                        pair = self.model.NewBoolVar(f'consec_r_v{v}_d{d1}_{d2}')
                        self.model.AddMultiplicationEquality(pair, [
                            self.rappelable[(v, d1)], 
                            self.rappelable[(v, d2)]
                        ])
                        consecutive_pairs.append(pair)
                
                if consecutive_pairs:
                    self.model.Add(sum(consecutive_pairs) <= 1)
        
        # 7. Respect des jours de repos hebdomadaires
        print("✓ Contrainte 7: Respect des jours de repos hebdomadaires")
        for v in range(self.n_vets):
            vet_info = self.veterinaires[self.vet_names[v]]
            vet_name = self.vet_names[v]
            
            # Gérer jour_repos comme int ou liste
            jours_repos = vet_info['jour_repos']
            if isinstance(jours_repos, int):
                jours_repos = [jours_repos]
            
            conges = [datetime.strptime(c, '%Y-%m-%d') for c in vet_info.get('conges', [])]
            
            # NOUVEAU: Contraintes spéciales pour Dr. Olivier
            if vet_name == 'Dr. Olivier':
                for d in range(self.n_days):
                    # Dr. Olivier ne fait JAMAIS de garde premier
                    self.model.Add(self.premier[(v, d)] == 0)
                    # Dr. Olivier ne fait JAMAIS de 2ème de garde (week-end)
                    self.model.Add(self.deuxieme[(v, d)] == 0)
                    # Dr. Olivier ne fait PAS de rappelable les week-ends
                    if self.dates[d].weekday() >= 5:
                        self.model.Add(self.rappelable[(v, d)] == 0)
            if vet_name == 'Dr. Dorra':
                for d in range(self.n_days):
                    # Dr. Dorra ne fait JAMAIS de garde le lundi
                    if self.dates[d].weekday() == 0:  # Lundi
                        self.model.Add(self.premier[(v, d)] == 0)
                        self.model.Add(self.rappelable[(v, d)] == 0)
                        self.model.Add(self.deuxieme[(v, d)] == 0)
            if vet_name == 'Dr. Isaure':
                for d in range(self.n_days):
                    # Dr. Isaure ne fait JAMAIS de garde le lundi
                    if self.dates[d].weekday() == 0:  # Lundi
                        self.model.Add(self.premier[(v, d)] == 0)
                        self.model.Add(self.rappelable[(v, d)] == 0)
                        self.model.Add(self.deuxieme[(v, d)] == 0)
            
            for d in range(self.n_days):
                date = self.dates[d]
                
                # Pas de garde pendant les jours de repos hebdomadaires (seulement jours de semaine)
                if date.weekday() in jours_repos and date.weekday() < 5:
                    self.model.Add(self.premier[(v, d)] == 0)
                    self.model.Add(self.rappelable[(v, d)] == 0)
                
                # Pas de garde pendant les congés
                if date in conges:
                    self.model.Add(self.premier[(v, d)] == 0)
                    self.model.Add(self.rappelable[(v, d)] == 0)
                    self.model.Add(self.deuxieme[(v, d)] == 0)
                
                # Pas de garde PREMIER la veille d'un jour de repos ou congé
                # EXCEPTION: Dr. Julien peut être de garde la veille de ses jours de repos
                if d + 1 < self.n_days:
                    next_date = self.dates[d + 1]
                    
                    if vet_name == 'Dr. Julien':
                        # Seulement bloquer la veille des congés
                        if next_date in conges:
                            self.model.Add(self.premier[(v, d)] == 0)
                    elif vet_name != 'Dr. Olivier':  # Olivier ne fait jamais de premier de toute façon
                        # Pour les autres : bloquer la veille des jours de repos ET des congés
                        if (next_date.weekday() in jours_repos and next_date.weekday() < 5) or next_date in conges:
                            self.model.Add(self.premier[(v, d)] == 0)
                            self.model.Add(self.rappelable[(v, d)] == 0)
        
        # 9. Pas de garde vendredi si garde week-end (QUEL QUE SOIT LE RÔLE)
        print("✓ Contrainte 9: Pas vendredi si week-end")
        for v in range(self.n_vets):
            for weekend_days in self.weekends:
                if len(weekend_days) > 0:
                    sam = weekend_days[0]
                    if sam > 0 and self.dates[sam - 1].weekday() == 4:
                        vendredi = sam - 1
                        
                        # Interdire TOUTES les combinaisons vendredi + week-end
                        # Si garde vendredi (premier OU rappelable), alors PAS de garde week-end
                        self.model.Add(self.premier[(v, vendredi)] + self.premier[(v, sam)] <= 1)
                        self.model.Add(self.premier[(v, vendredi)] + self.deuxieme[(v, sam)] <= 1)
                        self.model.Add(self.rappelable[(v, vendredi)] + self.premier[(v, sam)] <= 1)
                        self.model.Add(self.rappelable[(v, vendredi)] + self.deuxieme[(v, sam)] <= 1)
        
        # 10. Espacement de 14 jours entre week-ends (pour 1er ET 2ème de garde)
        print("✓ Contrainte 10: Espacement 14 jours entre week-ends")
        for v in range(self.n_vets):
            for i, weekend_i in enumerate(self.weekends):
                if len(weekend_i) == 2:
                    sam_i = weekend_i[0]
                    for j, weekend_j in enumerate(self.weekends[i+1:], start=i+1):
                        if len(weekend_j) == 2:
                            sam_j = weekend_j[0]
                            days_apart = (self.dates[sam_j] - self.dates[sam_i]).days
                            
                            if days_apart < 14:
                                self.model.Add(
                                    self.premier[(v, sam_i)] + self.deuxieme[(v, sam_i)] +
                                    self.premier[(v, sam_j)] + self.deuxieme[(v, sam_j)] <= 1
                                )
        
        # 11. Équilibrage des gardes premier (écart max 2) - Dr. Olivier exclu
        # AVEC PRISE EN COMPTE DE L'HISTORIQUE
        print("✓ Contrainte 11: Équilibrage gardes premier (écart max 2) + historique")
        gardes_premier = []
        gardes_premier_vars = []
        
        for v in range(self.n_vets):
            if self.vet_names[v] not in ['Dr. Olivier', 'Dr. Dorra']:
                nb = self.model.NewIntVar(0, self.n_days, f'nb_premier_v{v}')
                self.model.Add(nb == sum(self.premier[(v, d)] for d in range(self.n_days)))
                gardes_premier_vars.append(nb)
                
                # Ajouter l'historique
                hist_count = (self.historical_stats[self.vet_names[v]]['premier_semaine'] + 
                             self.historical_stats[self.vet_names[v]]['premier_weekend'])
                
                gardes_premier.append(hist_count)  # Gardes passées
        
        if len(gardes_premier_vars) > 1:
            # Calculer le total (historique + nouveau planning) pour chaque vétérinaire
            totaux_premier = []
            for idx, var in enumerate(gardes_premier_vars):
                total = self.model.NewIntVar(0, self.n_days * 10, f'total_premier_{idx}')
                self.model.Add(total == var + gardes_premier[idx])
                totaux_premier.append(total)
            
            # Équilibrer les totaux
            min_p = self.model.NewIntVar(0, self.n_days * 10, 'min_premier_total')
            max_p = self.model.NewIntVar(0, self.n_days * 10, 'max_premier_total')
            self.model.AddMinEquality(min_p, totaux_premier)
            self.model.AddMaxEquality(max_p, totaux_premier)
            self.model.Add(max_p - min_p <= self.config.contraintes['ecart_equilibrage_premier'])
        
        # 12. Équilibrage rappelable EN SEMAINE (écart max 2) - Dr. Olivier exclu
        # AVEC PRISE EN COMPTE DE L'HISTORIQUE
        print("✓ Contrainte 12: Équilibrage rappelable EN SEMAINE (écart max 2) + historique")
        gardes_rappelable = []
        gardes_rappelable_vars = []
        
        for v in range(self.n_vets):
            if self.vet_names[v] not in ['Dr. Olivier', 'Dr. Dorra']:
                nb = self.model.NewIntVar(0, len(self.weekdays), f'nb_rappelable_v{v}')
                self.model.Add(nb == sum(self.rappelable[(v, d)] for d in self.weekdays))
                gardes_rappelable_vars.append(nb)
                
                # Ajouter l'historique
                hist_count = self.historical_stats[self.vet_names[v]]['rappelable_semaine']
                gardes_rappelable.append(hist_count)
        
        if len(gardes_rappelable_vars) > 1:
            # Calculer le total pour chaque vétérinaire
            totaux_rappelable = []
            for idx, var in enumerate(gardes_rappelable_vars):
                total = self.model.NewIntVar(0, len(self.weekdays) * 10, f'total_rappelable_{idx}')
                self.model.Add(total == var + gardes_rappelable[idx])
                totaux_rappelable.append(total)
            
            # Équilibrer les totaux
            min_r = self.model.NewIntVar(0, len(self.weekdays) * 10, 'min_rappelable_total')
            max_r = self.model.NewIntVar(0, len(self.weekdays) * 10, 'max_rappelable_total')
            self.model.AddMinEquality(min_r, totaux_rappelable)
            self.model.AddMaxEquality(max_r, totaux_rappelable)
            self.model.Add(max_r - min_r <= self.config.contraintes['ecart_equilibrage_rappelable'])
        
        # 13. NOUVEAU: Équilibrage 2ème de garde WEEK-END (écart max 2) - Dr. Olivier exclu
        # AVEC PRISE EN COMPTE DE L'HISTORIQUE
        print("✓ Contrainte 13: Équilibrage 2ème de garde WEEK-END (écart max 2) + historique")
        gardes_deuxieme = []
        gardes_deuxieme_vars = []
        
        for v in range(self.n_vets):
            if self.vet_names[v] not in ['Dr. Olivier', 'Dr. Dorra']:
                nb = self.model.NewIntVar(0, len(self.weekends) * 2, f'nb_deuxieme_v{v}')
                weekend_days = [d for weekend in self.weekends for d in weekend]
                self.model.Add(nb == sum(self.deuxieme[(v, d)] for d in weekend_days))
                gardes_deuxieme_vars.append(nb)
                
                # Ajouter l'historique
                hist_count = self.historical_stats[self.vet_names[v]]['deuxieme_weekend']
                gardes_deuxieme.append(hist_count)
        
        if len(gardes_deuxieme_vars) > 1:
            # Calculer le total pour chaque vétérinaire
            totaux_deuxieme = []
            for idx, var in enumerate(gardes_deuxieme_vars):
                total = self.model.NewIntVar(0, len(self.weekends) * 20, f'total_deuxieme_{idx}')
                self.model.Add(total == var + gardes_deuxieme[idx])
                totaux_deuxieme.append(total)
            
            # Équilibrer les totaux
            min_d = self.model.NewIntVar(0, len(self.weekends) * 20, 'min_deuxieme_total')
            max_d = self.model.NewIntVar(0, len(self.weekends) * 20, 'max_deuxieme_total')
            self.model.AddMinEquality(min_d, totaux_deuxieme)
            self.model.AddMaxEquality(max_d, totaux_deuxieme)
            self.model.Add(max_d - min_d <= self.config.contraintes['ecart_equilibrage_deuxieme'])
        
        # 14. NOUVEAU: Règle de compatibilité entre 1er de garde et rappelable/2ème
        print("✓ Contrainte 14: Règle de compatibilité des binômes")
        
        # Utiliser la configuration
        indices_A = [i for i, name in enumerate(self.vet_names) if name in self.config.groupe_A]
        indices_B = [i for i, name in enumerate(self.vet_names) 
             if name in self.config.groupe_B and name != 'Dr. Dorra']
        
        # Pour chaque jour (semaine et week-end)
        for d in range(self.n_days):
            is_weekend = self.dates[d].weekday() >= 5
            
            # Pour chaque vétérinaire du groupe A
            for v_a in indices_A:
                if is_weekend:
                    # Si groupe A est 1er de garde le week-end, alors 2ème doit être du groupe B
                    for v_autre in range(self.n_vets):
                        if v_autre not in indices_B and v_autre != v_a and self.vet_names[v_autre] != 'Dr. Dorra':
                            # Interdire ce binôme
                            self.model.Add(self.premier[(v_a, d)] + self.deuxieme[(v_autre, d)] <= 1)
                else:
                    # Si groupe A est 1er de garde en semaine, alors rappelable doit être du groupe B
                    for v_autre in range(self.n_vets):
                        if v_autre not in indices_B and v_autre != v_a and self.vet_names[v_autre] != 'Dr. Dorra':
                            # Interdire ce binôme
                            self.model.Add(self.premier[(v_a, d)] + self.rappelable[(v_autre, d)] <= 1)
        
        # 15. NOUVEAU: Repos le lundi après 2ème de garde le week-end (SAUF Dr. Laura si lundi = jour de repos)
        print("✓ Contrainte 15: Repos le lundi après 2ème de garde week-end (exception Dr. Laura)")
        for v in range(self.n_vets):
            vet_name = self.vet_names[v]
            vet_info = self.veterinaires[vet_name]
            
            # Récupérer les jours de repos
            jours_repos = vet_info['jour_repos']
            if isinstance(jours_repos, int):
                jours_repos = [jours_repos]
            
            # Récupérer les congés
            conges = [datetime.strptime(c, '%Y-%m-%d') for c in vet_info.get('conges', [])]
            
            for weekend_days in self.weekends:
                if len(weekend_days) == 2:
                    dim = weekend_days[1]  # Dimanche
                    if dim + 1 < self.n_days and self.dates[dim + 1].weekday() == 0:
                        lundi = dim + 1
                        lundi_date = self.dates[lundi]
                        
                        # EXCEPTION pour Dr. Laura: si le lundi est son jour de repos ET qu'elle n'est pas en congé
                        if vet_name == 'Dr. Laura' and 0 in jours_repos and lundi_date not in conges:
                            # Dr. Laura peut faire 2ème garde le week-end même si elle a repos le lundi
                            # On ne bloque PAS la 2ème garde week-end
                            pass
                        else:
                            # Pour tous les autres (ou Dr. Laura en congé le lundi)
                            # Si 2ème de garde ce week-end, repos le lundi
                            self.model.AddImplication(
                                self.deuxieme[(v, weekend_days[0])],
                                self.premier[(v, lundi)].Not()
                            )
                            self.model.AddImplication(
                                self.deuxieme[(v, weekend_days[0])],
                                self.rappelable[(v, lundi)].Not()
                            )
        
        print("--- FIN DES CONTRAINTES ---\n")
    
    def solve(self, time_limit=300):
        """Résout le problème"""
        self.add_constraints()
        
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = time_limit
        solver.parameters.num_search_workers = 8
        
        print(f"Recherche d'une solution (limite: {time_limit}s)...")
        status = solver.Solve(self.model)
        
        print(f"\nStatut: {solver.StatusName(status)}")
        
        if status == cp_model.OPTIMAL:
            print("✓ Solution optimale trouvée!")
            solution = self.extract_solution(solver)
            # Diagnostic automatique
            self.diagnose_schedule(solution)
            return solution
        elif status == cp_model.FEASIBLE:
            print("✓ Solution faisable trouvée!")
            solution = self.extract_solution(solver)
            # Diagnostic automatique
            self.diagnose_schedule(solution)
            return solution
        else:
            print(f"✗ Aucune solution trouvée: {solver.StatusName(status)}")
            return None
    
    def extract_solution(self, solver):
        """Extrait la solution"""
        schedule = []
        
        for d in range(self.n_days):
            date = self.dates[d]
            premier_vet = None
            rappelable_vet = None
            deuxieme_vet = None
            
            for v in range(self.n_vets):
                if solver.Value(self.premier[(v, d)]):
                    premier_vet = self.vet_names[v]
                if solver.Value(self.rappelable[(v, d)]):
                    rappelable_vet = self.vet_names[v]
                if solver.Value(self.deuxieme[(v, d)]):
                    deuxieme_vet = self.vet_names[v]
            
            schedule.append({
                'date': date.strftime('%Y-%m-%d'),
                'jour': date.strftime('%A'),
                'premier': premier_vet,
                'rappelable': rappelable_vet,
                'deuxieme': deuxieme_vet
            })
        
        return schedule
    
    def print_schedule(self, schedule):
        """Affiche le planning"""
        if schedule is None:
            print("Aucune solution trouvée!")
            return
        
        print("\n" + "="*80)
        print("PLANNING DE GARDES VÉTÉRINAIRES")
        print("="*80)
        
        for entry in schedule:
            date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
            jour_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche'][date_obj.weekday()]
            is_weekend = date_obj.weekday() >= 5
            
            print(f"\n{entry['date']} ({jour_fr})")
            print(f"  Premier de garde: {entry['premier']}")
            if is_weekend and entry['deuxieme']:
                print(f"  2ème de garde: {entry['deuxieme']}")
            elif entry['rappelable']:
                print(f"  Rappelable: {entry['rappelable']}")
    
    def export_to_excel(self, schedule, filename='planning_gardes.xlsx'):
        """Exporte le planning vers Excel avec mise en forme et couleurs par vétérinaire"""
        if schedule is None:
            print("Aucune solution à exporter!")
            return
        
        # Définir une palette de couleurs pour chaque vétérinaire
        colors = [
            'FFB3BA', 'FFDFBA', 'FFFFBA', 'BAFFC9', 'BAE1FF', 
            'E0BBE4', 'FFDFD3', 'D4F1F4', 'FFE5B4', 'C9E4DE',
            'F0E68C', 'DDA0DD', 'F5DEB3',
        ]
        
        # Créer un mapping vétérinaire -> couleur
        vet_colors = {}
        for idx, vet_name in enumerate(self.vet_names):
            vet_colors[vet_name] = colors[idx % len(colors)]
        
        data = []
        for entry in schedule:
            date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
            jour_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche'][date_obj.weekday()]
            is_weekend = date_obj.weekday() >= 5
            
            # Utiliser 2ème de garde pour les week-ends, rappelable pour la semaine
            second_role = entry['deuxieme'] if (is_weekend and entry['deuxieme']) else (entry['rappelable'] if entry['rappelable'] else '')
            
            # Vétérinaires en congé ce jour
            vets_en_conges = set()
            for vet_name, vet_info in self.veterinaires.items():
                conges = [datetime.strptime(c, '%Y-%m-%d') for c in vet_info.get('conges', [])]
                if date_obj in conges:
                    vets_en_conges.add(vet_name)
            
            # Vétérinaires en repos hebdomadaire ce jour
            vets_en_repos = set()
            for vet_name, vet_info in self.veterinaires.items():
                jours_repos = vet_info['jour_repos']
                if isinstance(jours_repos, int):
                    jours_repos = [jours_repos]
                if date_obj.weekday() in jours_repos and date_obj.weekday() < 5:  # Seulement jours de semaine
                    vets_en_repos.add(vet_name)
            
            # Vétérinaires présents à la clinique = tous - (congés + repos)
            tous_vets = set(self.vet_names)
            vets_presents = tous_vets - vets_en_conges - vets_en_repos
            vets_presents_str = ', '.join(sorted(vets_presents))
            nb_vets_presents = len(vets_presents)-1
            
            data.append({
                'Date': entry['date'],
                'Jour': jour_fr,
                'Premier de garde': entry['premier'] if entry['premier'] else '',
                'Rappelable/2ème': second_role,
                'Vétérinaires présents': vets_presents_str,
                'Nombre présents (après-midi)': nb_vets_presents  # AJOUT
            })
        
        df = pd.DataFrame(data)
        
        # Créer le fichier Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Planning', index=False)
        
        # Charger le fichier pour appliquer les styles
        wb = load_workbook(filename)
        ws = wb['Planning']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        weekend_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        border_style = Side(style='thin', color='000000')
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # Appliquer le style aux en-têtes
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Appliquer le style aux lignes de données
        for row_num in range(2, ws.max_row + 1):
            jour = ws.cell(row=row_num, column=2).value
            premier_vet = ws.cell(row=row_num, column=3).value
            second_vet = ws.cell(row=row_num, column=4).value
            
            # Colonnes Date et Jour
            for col in [1, 2]:
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if jour in ['Samedi', 'Dimanche']:
                    cell.fill = weekend_fill
            
            # Colonne Premier de garde
            cell_premier = ws.cell(row=row_num, column=3)
            cell_premier.border = border
            cell_premier.alignment = Alignment(horizontal='left', vertical='center')
            if premier_vet and premier_vet in vet_colors:
                color = vet_colors[premier_vet]
                cell_premier.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell_premier.font = Font(bold=True, size=11)
            
            # Colonne Rappelable/2ème
            cell_second = ws.cell(row=row_num, column=4)
            cell_second.border = border
            cell_second.alignment = Alignment(horizontal='left', vertical='center')
            if second_vet and second_vet in vet_colors:
                color = vet_colors[second_vet]
                cell_second.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell_second.font = Font(size=11)
            
            # Colonne Vétérinaires présents
            cell_presents = ws.cell(row=row_num, column=5)
            cell_presents.border = border
            cell_presents.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell_presents.font = Font(size=10)

            # Colonne Nombre présents
            cell_nb_presents = ws.cell(row=row_num, column=6)
            cell_nb_presents.border = border
            cell_nb_presents.alignment = Alignment(horizontal='center', vertical='center')
            cell_nb_presents.font = Font(size=11, bold=True)
            
            # Remplissage rouge si < 6
            nb_presents = cell_nb_presents.value
            if nb_presents and nb_presents < 6:
                cell_nb_presents.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                cell_nb_presents.font = Font(size=11, bold=True, color='FFFFFF')
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 50  # Colonne large pour la liste des vétérinaires
        ws.column_dimensions['F'].width = 25  # AJOUT : Colonne Nombre présents
        
        # Sauvegarder avant d'ajouter les statistiques
        wb.save(filename)
        
        # Ajouter les statistiques
        stats_data = self.generate_statistics(schedule)
        df_stats = pd.DataFrame(stats_data)
        
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_stats.to_excel(writer, sheet_name='Statistiques', index=False)
        
        # Recharger pour styler les statistiques
        wb = load_workbook(filename)
        ws_stats = wb['Statistiques']
        
        # En-têtes statistiques
        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Lignes statistiques avec couleurs
        for row_num in range(2, ws_stats.max_row + 1):
            vet_name = ws_stats.cell(row=row_num, column=1).value
            
            for col_num in range(1, 6):  # 5 colonnes
                cell = ws_stats.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colorer la colonne du nom du vétérinaire
                if col_num == 1 and vet_name and vet_name in vet_colors:
                    color = vet_colors[vet_name]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(bold=True, size=11)
        
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 18
        ws_stats.column_dimensions['C'].width = 18
        ws_stats.column_dimensions['D'].width = 20
        ws_stats.column_dimensions['E'].width = 20
        
        # Sauvegarder le fichier final
        wb.save(filename)
        wb.close()
        
        print(f"\n✓ Planning exporté vers: {filename}")
        print("   Chaque vétérinaire a une couleur distinctive pour faciliter la lecture!")
    
    def generate_statistics(self, schedule):
        """Génère des statistiques par vétérinaire"""
        stats = defaultdict(lambda: {
            'premier_semaine': 0, 
            'premier_weekend': 0, 
            'rappelable_semaine': 0,
            'deuxieme_weekend': 0
        })
        
        for entry in schedule:
            date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
            is_weekend = date_obj.weekday() >= 5
            
            if entry['premier']:
                if is_weekend:
                    stats[entry['premier']]['premier_weekend'] += 1
                else:
                    stats[entry['premier']]['premier_semaine'] += 1
            
            if entry['rappelable']:
                stats[entry['rappelable']]['rappelable_semaine'] += 1
            
            if entry['deuxieme']:
                stats[entry['deuxieme']]['deuxieme_weekend'] += 1
        
        stats_list = []
        for vet_name in sorted(stats.keys()):
            vet_stats = stats[vet_name]
            nb_weekends_premier = vet_stats['premier_weekend'] // 2
            nb_weekends_deuxieme = vet_stats['deuxieme_weekend'] // 2
            
            stats_list.append({
                'Vétérinaire': vet_name,
                'Premier semaine': vet_stats['premier_semaine'],
                'Premier WE': nb_weekends_premier,
                'Rappelable semaine': vet_stats['rappelable_semaine'],
                '2ème de garde WE': nb_weekends_deuxieme
            })
        
        return stats_list


# EXEMPLE D'UTILISATION
if __name__ == "__main__":
    # Configuration officielle des vétérinaires
    veterinaires = {
        'Dr. Julien': {'jour_repos': [1, 3], 'conges': []},    # Mardi et jeudi
        'Dr. Maxime': {'jour_repos': 2, 'conges': []},         # Mercredi
        'Dr. Isaure': {'jour_repos': 4, 'conges': []},         # Vendredi
        'Dr. Mélanie': {'jour_repos': 2, 'conges': []},        # Mercredi
        'Dr. Nicolas': {'jour_repos': 3, 'conges': []},        # Jeudi
        'Dr. Timoty': {'jour_repos': [3, 4], 'conges': []},    # Jeudi et vendredi
        'Dr. Laura': {'jour_repos': 0, 'conges': []},          # Lundi
        'Dr. Lauranne': {'jour_repos': 4, 'conges': []},       # Vendredi
        'Dr. Malaurie': {'jour_repos': 1, 'conges': []},       # Mardi
        'Dr. Sarah': {'jour_repos': 2, 'conges': []},          # Mercredi
        'Dr. Olivier': {'jour_repos': [2], 'conges': []},       # Mercredi
        'Dr. Dorra': {'jour_repos': [1, 2, 3], 'conges': []},       # Mardi, Mercredi et Jeudi
    }
    
    # Charger ou créer l'historique
    history = VetSchedulerHistory('historique_gardes.json')
    
    # Afficher l'historique existant
    if history.history:
        print("\n📚 HISTORIQUE CHARGÉ:")
        history.print_history()
    
    # Génération du planning pour 5 semaines à partir du 5 janvier 2026
    print("\n" + "="*80)
    print("GÉNÉRATION DU PLANNING - 5 SEMAINES")
    print("="*80)
    print("\n💡 EXCEPTION ACTIVÉE: Dr. Laura peut être de garde le week-end")
    print("   même si elle a son repos le lundi (sauf si elle est en congé)")
    
    # Date de début : lundi 5 janvier 2026
    # Date de fin : dimanche 8 février 2026 (5 semaines complètes)
    scheduler = VetScheduler('2026-01-05', '2026-02-08', veterinaires, history=history)
    schedule = scheduler.solve(time_limit=300)
    
    if schedule:
        scheduler.print_schedule(schedule)
        scheduler.export_to_excel(schedule, 'planning_janvier_2026.xlsx')
        
        # Ajouter ce planning à l'historique
        history.add_schedule(schedule, period_name='janvier_2026')
        
        print("\n" + "="*80)
        print("🎉 PLANNING GÉNÉRÉ AVEC SUCCÈS !")
        print("="*80)
        print("\n💡 L'historique a été mis à jour. Le prochain planning tiendra compte")
        print("   de l'équité sur plusieurs mois.")
    else:
        print("\n❌ Impossible de générer le planning")